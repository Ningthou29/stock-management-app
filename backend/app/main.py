from fastapi import FastAPI, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import List, Optional
import logging
import io
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from .database import supabase
from .schemas import (
    EquipmentCreate,
    EquipmentUpdate,
    EquipmentResponse,
    SaleCreate,
    DashboardMetrics
)

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Cricket Equipment Stock Maintenance API",
    description="Backend service for tracking cricket inventory and sales",
    version="1.0.0"
)

# Enable CORS for frontend integration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust for production security if needed
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Additional Schemas ---
class SaleResponse(BaseModel):
    id: str
    equipment_name: str
    category: str
    quantity_sold: int
    sale_price: float
    total_revenue: float
    profit: float
    created_at: str
    stock_arrival_date: str

class SalesExportItem(BaseModel):
    equipment_name: str
    category: str
    quantity_sold: int
    sale_price: float
    total_revenue: float
    profit: float
    stock_arrival_date: str
    sale_date: str
    sale_time: str

@app.get("/")
def read_root():
    return {"message": "Cricket Equipment Stock Maintenance API is active."}

@app.get("/api/dashboard/metrics", response_model=DashboardMetrics)
def get_dashboard_metrics():
    try:
        response = supabase.table("equipment").select("*").execute()
        items = response.data or []
        
        total_investment = sum(item["current_stock"] * float(item["cost_price"]) for item in items)
        potential_revenue = sum(item["current_stock"] * float(item["selling_price"]) for item in items)
        potential_profit = potential_revenue - total_investment
        
        low_stock_alerts = [
            item for item in items 
            if item["current_stock"] <= item["min_stock_threshold"]
        ]
        
        total_unique_items = len(items)
        total_stock_count = sum(item["current_stock"] for item in items)
        
        return {
            "total_investment": total_investment,
            "potential_revenue": potential_revenue,
            "potential_profit": potential_profit,
            "low_stock_count": len(low_stock_alerts),
            "low_stock_alerts": low_stock_alerts,
            "total_unique_items": total_unique_items,
            "total_stock_count": total_stock_count
        }
    except Exception as e:
        logger.error(f"Error fetching dashboard metrics: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.get("/api/inventory", response_model=List[EquipmentResponse])
def get_inventory():
    try:
        # Fetch inventory, sort alphabetically by name
        response = supabase.table("equipment").select("*").order("name").execute()
        return response.data or []
    except Exception as e:
        logger.error(f"Error fetching inventory: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.get("/api/inventory/export")
def export_inventory_excel():
    """Export the full inventory catalog as a styled .xlsx file with updated columns."""
    try:
        response = supabase.table("equipment").select("*").order("name").execute()
        items = response.data or []

        sales_res = supabase.table("sales").select("*").execute()
        sales = sales_res.data or []
    except Exception as e:
        logger.error(f"Error fetching inventory for export: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Catalog"

    # --- Styles ---
    header_fill   = PatternFill("solid", fgColor="0A251C")   # dark cricket-pitch green
    header_font   = Font(bold=True, color="C5A85A", name="Calibri", size=11)
    alt_row_fill  = PatternFill("solid", fgColor="F1F5F9")   # light slate
    border_side   = Side(style="thin", color="CBD5E1")
    cell_border   = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )
    center_align  = Alignment(horizontal="center", vertical="center")
    left_align    = Alignment(horizontal="left",   vertical="center")

    display_headers = [
        "Item", "Category", "Total Quantity",
        "Cost Per Unit (₹)", "Total Cost (₹)", "Item Entry Date", "Item Exit Date", "Item Exit Quantity"
    ]

    # Write header row
    for col_num, header_text in enumerate(display_headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = cell_border

    # Write data rows
    for row_num, item in enumerate(items, start=2):
        fill = alt_row_fill if row_num % 2 == 0 else None

        eq_id = item.get("id")
        eq_name = item.get("name")
        item_sales = [s for s in sales if s.get("equipment_id") == eq_id or s.get("equipment_name") == eq_name]
        total_exit_qty = sum(s.get("quantity_sold", 0) for s in item_sales)

        sorted_sales = sorted(
            item_sales,
            key=lambda x: str(x.get("created_at") or x.get("sold_at") or ""),
            reverse=True
        )
        latest_sale = sorted_sales[0] if sorted_sales else None
        exit_date_str = str(latest_sale.get("created_at") or latest_sale.get("sold_at") or "N/A") if latest_sale else "N/A"

        current_stock = item.get("current_stock", 0)
        cost_price = float(item.get("cost_price", 0.0))
        total_cost = current_stock * cost_price

        row_values = [
            item.get("name", ""),
            item.get("category", ""),
            current_stock,
            cost_price,
            total_cost,
            item.get("created_at", ""),
            exit_date_str,
            total_exit_qty
        ]

        for col_num, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border    = cell_border
            cell.alignment = center_align if col_num in [3, 6, 7, 8] else (left_align if col_num <= 2 else center_align)
            if fill:
                cell.fill = fill

    # Auto-size columns
    col_widths = [30, 16, 16, 18, 18, 24, 24, 18]
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 22

    # Stream as HTTP response
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_catalog.xlsx"}
    )

@app.post("/api/inventory", response_model=EquipmentResponse, status_code=status.HTTP_201_CREATED)
def add_inventory_item(item: EquipmentCreate):
    try:
        payload = item.model_dump()
        response = supabase.table("equipment").insert(payload).execute()
        if not response.data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to create equipment item"
            )
        return response.data[0]
    except Exception as e:
        logger.error(f"Error adding inventory item: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.put("/api/inventory/{id}", response_model=EquipmentResponse)
def update_inventory_item(id: str, item: EquipmentUpdate):
    try:
        # Remove None fields to avoid overwriting existing properties with null
        payload = {k: v for k, v in item.model_dump().items() if v is not None}
        if not payload:
            # Nothing to update, retrieve original
            response = supabase.table("equipment").select("*").eq("id", id).execute()
            if not response.data:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
            return response.data[0]
            
        response = supabase.table("equipment").update(payload).eq("id", id).execute()
        if not response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Item not found or failed to update"
            )
        return response.data[0]
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error updating inventory item {id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.delete("/api/inventory/{id}")
def delete_inventory_item(id: str):
    try:
        response = supabase.table("equipment").delete().eq("id", id).execute()
        if not response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Item not found or already deleted"
            )
        return {"message": "Item deleted successfully", "id": id}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error deleting inventory item {id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.post("/api/sales")
def log_sale(sale: SaleCreate):
    try:
        # Call the PostgreSQL RPC stored procedure to safely execute stock decrement
        # and sale logging in a transaction.
        response = supabase.rpc("log_sale", {
            "p_equipment_id": sale.equipment_id,
            "p_quantity_sold": sale.quantity_sold,
            "p_sale_price": sale.sale_price
        }).execute()
        
        # response.data returns the new stock levels
        new_stock = response.data
        
        return {
            "message": "Sale logged successfully",
            "equipment_id": sale.equipment_id,
            "quantity_sold": sale.quantity_sold,
            "sale_price": sale.sale_price,
            "new_stock": new_stock
        }
    except Exception as e:
        err_msg = str(e)
        logger.error(f"Error logging sale: {err_msg}")
        
        # Check database raised exceptions
        if "Insufficient stock" in err_msg:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=err_msg
            )
        elif "Equipment item not found" in err_msg:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Equipment item not found."
            )
            
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database transaction error: {err_msg}"
        )

@app.delete("/api/sales/{sale_id}")
def delete_sale_record(sale_id: str):
    """
    Delete a sale record and restore the stock quantity.
    This ensures data consistency by rolling back the stock update.
    """
    try:
        # First, get the sale details to know how much stock to restore
        sale_response = supabase.table("sales_log") \
            .select("equipment_id, quantity_sold") \
            .eq("id", sale_id) \
            .execute()
        
        if not sale_response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Sale record not found"
            )
        
        sale = sale_response.data[0]
        equipment_id = sale["equipment_id"]
        quantity_sold = sale["quantity_sold"]
        
        # Restore the stock quantity
        equipment_response = supabase.table("equipment") \
            .select("current_stock") \
            .eq("id", equipment_id) \
            .execute()
        
        if not equipment_response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Equipment item not found"
            )
        
        current_stock = equipment_response.data[0]["current_stock"]
        new_stock = current_stock + quantity_sold
        
        # Update equipment stock
        supabase.table("equipment") \
            .update({"current_stock": new_stock}) \
            .eq("id", equipment_id) \
            .execute()
        
        # Delete the sale record
        delete_response = supabase.table("sales_log") \
            .delete() \
            .eq("id", sale_id) \
            .execute()
        
        if not delete_response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Sale record not found or already deleted"
            )
        
        return {
            "message": "Sale record deleted successfully",
            "id": sale_id,
            "restored_stock": new_stock,
            "quantity_restored": quantity_sold
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error deleting sale record {sale_id}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.get("/api/sales", response_model=List[SaleResponse])
def get_sales_log():
    """Get all sales records with equipment details."""
    try:
        # Get all sales with equipment data
        response = supabase.table("sales_log") \
            .select("*, equipment(id, name, category, cost_price, selling_price, created_at)") \
            .order("sold_at", desc=True) \
            .execute()
        
        if hasattr(response, 'error') and response.error:
            logger.error(f"Supabase error: {response.error}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(response.error)
            )
        
        # Format the data for frontend
        sales_list = []
        for row in response.data or []:
            eq_data = row.get("equipment", {})
            
            # Calculate total revenue and profit
            quantity = row.get("quantity_sold", 0)
            sale_price = float(row.get("sale_price", 0))
            cost_price = float(eq_data.get("cost_price", 0))
            total_revenue = quantity * sale_price
            profit = total_revenue - (quantity * cost_price)
            
            sales_list.append({
                "id": row.get("id"),
                "equipment_name": eq_data.get("name", "Unknown Item"),
                "category": eq_data.get("category", "Uncategorized"),
                "quantity_sold": quantity,
                "sale_price": sale_price,
                "total_revenue": total_revenue,
                "profit": profit,
                "created_at": row.get("sold_at", datetime.now().isoformat()),
                "stock_arrival_date": eq_data.get("created_at", datetime.now().isoformat())
            })
        
        return sales_list
    except Exception as e:
        logger.error(f"Error fetching sales log: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Database error: {str(e)}"
        )

@app.post("/api/sales/export")
def export_sales_excel(export_data: List[SalesExportItem]):
    """Export sales data as a styled .xlsx file."""
    try:
        if not export_data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No sales data to export"
            )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Log"

        # --- Styles ---
        header_fill   = PatternFill("solid", fgColor="0A251C")
        header_font   = Font(bold=True, color="C5A85A", name="Calibri", size=11)
        alt_row_fill  = PatternFill("solid", fgColor="F1F5F9")
        border_side   = Side(style="thin", color="CBD5E1")
        cell_border   = Border(
            left=border_side, right=border_side,
            top=border_side, bottom=border_side
        )
        center_align  = Alignment(horizontal="center", vertical="center")
        left_align    = Alignment(horizontal="left", vertical="center")
        
        # Currency format for numbers
        currency_format = "#,##0.00"

        display_headers = [
            "Item Name", "Category", "Quantity Sold",
            "Sale Price (₹)", "Total Revenue (₹)", "Profit (₹)",
            "Stock Arrival Date", "Sale Date", "Sale Time"
        ]
        field_names = [
            "equipment_name", "category", "quantity_sold",
            "sale_price", "total_revenue", "profit",
            "stock_arrival_date", "sale_date", "sale_time"
        ]

        # Write header row
        for col_num, header_text in enumerate(display_headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = cell_border

        # Write data rows
        for row_num, item in enumerate(export_data, start=2):
            fill = alt_row_fill if row_num % 2 == 0 else None
            for col_num, field in enumerate(field_names, start=1):
                value = getattr(item, field, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = cell_border
                
                # Apply alignment and formatting
                if field in ["quantity_sold", "sale_price", "total_revenue", "profit"]:
                    cell.alignment = center_align
                    if isinstance(value, (int, float)):
                        cell.number_format = currency_format
                else:
                    cell.alignment = left_align
                
                if fill:
                    cell.fill = fill

        # Auto-size columns
        col_widths = [30, 16, 16, 16, 18, 16, 20, 16, 16]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

        ws.row_dimensions[1].height = 22

        # Stream as HTTP response
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=sales_log_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            }
        )
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Error exporting sales: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Export error: {str(e)}"
        )

@app.post("/api/sales/upload")
async def upload_sales_excel(file: UploadFile = File(...)):
    """Import sales records from Excel file."""
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload a standard Excel file (.xlsx or .xls)."
        )

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read Excel file structure: {str(e)}"
        )

    # Extract headers
    try:
        first_row = next(ws.iter_rows(max_row=1))
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in first_row]
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded Excel sheet appears to be empty."
        )

    # Required columns for sales import
    required_cols = ["equipment_name", "quantity_sold", "sale_price", "sale_date"]
    for col in required_cols:
        if col not in headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required column '{col}'. Required headers: {', '.join(required_cols)}"
            )

    name_idx = headers.index("equipment_name")
    quantity_idx = headers.index("quantity_sold")
    price_idx = headers.index("sale_price")
    date_idx = headers.index("sale_date")

    imported_count = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully empty rows
        if not row or all(cell is None for cell in row):
            continue

        try:
            equipment_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            if not equipment_name:
                errors.append(f"Row {row_num}: Equipment name cannot be empty.")
                continue

            # Parse quantity
            try:
                quantity_sold = int(row[quantity_idx])
                if quantity_sold <= 0:
                    errors.append(f"Row {row_num} ('{equipment_name}'): Quantity must be positive.")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num} ('{equipment_name}'): Invalid quantity_sold value.")
                continue

            # Parse sale price
            try:
                sale_price = float(row[price_idx])
                if sale_price <= 0:
                    errors.append(f"Row {row_num} ('{equipment_name}'): Sale price must be positive.")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num} ('{equipment_name}'): Invalid sale_price value.")
                continue

            # Parse sale date
            try:
                if row[date_idx] is None:
                    errors.append(f"Row {row_num} ('{equipment_name}'): Sale date cannot be empty.")
                    continue
                
                if isinstance(row[date_idx], datetime):
                    sold_at = row[date_idx]
                else:
                    # Try parsing various date formats
                    date_str = str(row[date_idx]).strip()
                    # Try different formats
                    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            sold_at = datetime.strptime(date_str, fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        raise ValueError(f"Unsupported date format: {date_str}")
            except Exception as e:
                errors.append(f"Row {row_num} ('{equipment_name}'): Invalid sale_date format. Use YYYY-MM-DD or YYYY-MM-DD HH:MM:SS")
                continue

            # Find equipment by name
            eq_response = supabase.table("equipment").select("id, current_stock, cost_price").eq("name", equipment_name).execute()
            
            if not eq_response.data:
                errors.append(f"Row {row_num} ('{equipment_name}'): Equipment not found in inventory.")
                continue

            equipment = eq_response.data[0]
            equipment_id = equipment["id"]
            current_stock = equipment["current_stock"]
            
            # Check if enough stock
            if current_stock < quantity_sold:
                errors.append(f"Row {row_num} ('{equipment_name}'): Insufficient stock. Available: {current_stock}, Requested: {quantity_sold}")
                continue

            # Log the sale using the RPC function
            try:
                response = supabase.rpc("log_sale", {
                    "p_equipment_id": equipment_id,
                    "p_quantity_sold": quantity_sold,
                    "p_sale_price": sale_price,
                    "p_sold_at": sold_at.isoformat()
                }).execute()
                
                # Check if the RPC returned an error
                if hasattr(response, 'error') and response.error:
                    errors.append(f"Row {row_num} ('{equipment_name}'): {str(response.error)}")
                    continue
                    
                imported_count += 1
            except Exception as e:
                errors.append(f"Row {row_num} ('{equipment_name}'): Database error - {str(e)}")
                continue

        except Exception as e:
            errors.append(f"Row {row_num}: Unexpected error - {str(e)}")

    return {
        "message": "Sales import completed.",
        "imported_count": imported_count,
        "errors": errors
    }

@app.post("/api/inventory/upload")
async def upload_inventory_excel(file: UploadFile = File(...)):
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid file format. Please upload a standard Excel file (.xlsx or .xls)."
        )

    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read Excel file structure: {str(e)}"
        )

    # Extract headers
    try:
        first_row = next(ws.iter_rows(max_row=1))
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in first_row]
    except Exception as e:
         raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded Excel sheet appears to be empty."
        )

    required_cols = ["name", "category", "current_stock", "cost_price", "selling_price"]
    for col in required_cols:
        if col not in headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required column '{col}'. Excel sheet must have headers: name, category, current_stock, cost_price, selling_price"
            )

    name_idx = headers.index("name")
    category_idx = headers.index("category")
    stock_idx = headers.index("current_stock")
    cost_idx = headers.index("cost_price")
    sell_idx = headers.index("selling_price")
    threshold_idx = headers.index("min_stock_threshold") if "min_stock_threshold" in headers else -1

    VALID_CATEGORIES = {'Bats', 'Balls', 'Gloves', 'Pads', 'Helmets', 'Accessories', 'Bags', 'Clothing'}
    imported_count = 0
    updated_count = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip fully empty rows
        if not row or all(cell is None for cell in row):
            continue

        try:
            name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
            if not name:
                errors.append(f"Row {row_num}: Item name cannot be empty.")
                continue

            category = str(row[category_idx]).strip() if row[category_idx] is not None else ""
            category_normalized = category.capitalize()
            
            # Simple singular to plural conversions to help user imports
            if category_normalized not in VALID_CATEGORIES:
                if category_normalized == 'Bat':
                    category_normalized = 'Bats'
                elif category_normalized == 'Ball':
                    category_normalized = 'Balls'
                elif category_normalized == 'Glove':
                    category_normalized = 'Gloves'
                elif category_normalized == 'Pad':
                    category_normalized = 'Pads'
                elif category_normalized == 'Helmet':
                    category_normalized = 'Helmets'
                elif category_normalized == 'Bag':
                    category_normalized = 'Bags'
                elif category_normalized == 'Accessory':
                    category_normalized = 'Accessories'

            if category_normalized not in VALID_CATEGORIES:
                errors.append(f"Row {row_num} ('{name}'): Invalid category '{category}'. Must be one of {list(VALID_CATEGORIES)}.")
                continue

            # Parse stock
            try:
                current_stock = int(row[stock_idx])
                if current_stock < 0:
                    errors.append(f"Row {row_num} ('{name}'): Stock level cannot be negative.")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num} ('{name}'): Invalid current_stock value.")
                continue

            # Parse cost price
            try:
                cost_price = float(row[cost_idx])
                if cost_price < 0:
                    errors.append(f"Row {row_num} ('{name}'): Cost price cannot be negative.")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num} ('{name}'): Invalid cost_price value.")
                continue

            # Parse selling price
            try:
                selling_price = float(row[sell_idx])
                if selling_price < 0:
                    errors.append(f"Row {row_num} ('{name}'): Selling price cannot be negative.")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Row {row_num} ('{name}'): Invalid selling_price value.")
                continue

            # Parse min stock threshold
            min_stock_threshold = 5
            if threshold_idx != -1 and row[threshold_idx] is not None:
                try:
                    min_stock_threshold = int(row[threshold_idx])
                    if min_stock_threshold < 0:
                        min_stock_threshold = 5
                except (ValueError, TypeError):
                    pass

            payload = {
                "name": name,
                "category": category_normalized,
                "current_stock": current_stock,
                "min_stock_threshold": min_stock_threshold,
                "cost_price": cost_price,
                "selling_price": selling_price
            }

            # Check if item exists in db by name (exact check)
            check_res = supabase.table("equipment").select("id").eq("name", name).execute()

            if check_res.data:
                # Update existing
                item_id = check_res.data[0]["id"]
                supabase.table("equipment").update(payload).eq("id", item_id).execute()
                updated_count += 1
            else:
                # Insert new
                supabase.table("equipment").insert(payload).execute()
                imported_count += 1

        except Exception as e:
            errors.append(f"Row {row_num}: Database error during import: {str(e)}")

    return {
        "message": "Import completed.",
        "imported_count": imported_count,
        "updated_count": updated_count,
        "errors": errors
    }